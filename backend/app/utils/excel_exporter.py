import io
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    @staticmethod
    def generate_excel(results: dict) -> io.BytesIO:
        """
        Main entry point for generating Excel. Inspects results payload
        to determine if it is single-horizon or long-term simulation.
        Returns a BytesIO stream.
        """
        is_long_term = results.get("is_long_term", False)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if is_long_term:
                ExcelExporter._write_long_term(writer, results)
            else:
                ExcelExporter._write_single_horizon(writer, results)
                
        output.seek(0)
        return output

    @staticmethod
    def _write_single_horizon(writer, results: dict):
        # 1. Sheet 1: Summary
        ExcelExporter._create_single_summary_sheet(writer, results)
        
        # 2. Sheet 2: Depot Timeseries
        ExcelExporter._create_single_timeseries_sheet(writer, results)
        
        # 3. Sheet 3: Bus Diagnostics
        ExcelExporter._create_bus_diagnostics_sheet(writer, results)
        
        # 4. Sheet 4: Bus Schedules (kW)
        ExcelExporter._create_bus_schedules_sheet(writer, results)
        
        # 5. Sheet 5: Bus SoC Profiles (kWh)
        ExcelExporter._create_bus_soc_sheet(writer, results)

    @staticmethod
    def _write_long_term(writer, results: dict):
        # 1. Sheet 1: Summary
        ExcelExporter._create_lt_summary_sheet(writer, results)
        
        # 2. Sheet 2: Daily Aggregates
        ExcelExporter._create_lt_daily_aggregates_sheet(writer, results)
        
        # 3. Sheet 3: Chronological Timeseries
        ExcelExporter._create_lt_timeseries_sheet(writer, results)
        
        # 4. Sheet 4: Bus Performance Summary
        ExcelExporter._create_lt_bus_performance_sheet(writer, results)

    # --- Helper Styling Methods ---
    @staticmethod
    def _apply_table_formatting(writer, sheet_name, title=""):
        workbook = writer.book
        worksheet = workbook[sheet_name]
        worksheet.views.sheetView[0].showGridLines = True
        
        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep navy
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        # Title formatting if provided
        start_row = 1
        if title:
            worksheet.merge_cells("A1:D1")
            worksheet["A1"] = title
            worksheet["A1"].font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
            worksheet.row_dimensions[1].height = 35
            worksheet["A1"].alignment = Alignment(vertical="center")
            start_row = 3
            
        # Format table headers and data
        for row in worksheet.iter_rows(min_row=start_row):
            is_header = (row[0].row == start_row)
            for cell in row:
                if is_header:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                else:
                    cell.border = thin_border
                    # Auto alignment based on content type
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = right_align
                    elif cell.value in ["healthy", "Healthy"]:
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # light green
                        cell.font = Font(color="065F46", bold=True)
                        cell.alignment = center_align
                    elif cell.value in ["warning", "Warning"]:
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # light yellow
                        cell.font = Font(color="92400E", bold=True)
                        cell.alignment = center_align
                    elif cell.value in ["critical", "Critical"]:
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # light red
                        cell.font = Font(color="991B1B", bold=True)
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
        # Row Heights
        for r_idx in range(start_row, worksheet.max_row + 1):
            worksheet.row_dimensions[r_idx].height = 24 if r_idx == start_row else 20
            
        # Autofit columns
        for col in worksheet.columns:
            max_len = 0
            for cell in col:
                if cell.row < start_row: continue # skip title merge
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    @staticmethod
    def _create_kpi_card(worksheet, start_col, start_row, label, value, unit="", color="1E3A8A"):
        """
        Creates a nice KPI block inside the summary tab.
        """
        label_cell = worksheet.cell(row=start_row, column=start_col)
        value_cell = worksheet.cell(row=start_row + 1, column=start_col)
        
        label_cell.value = label.upper()
        label_cell.font = Font(name="Calibri", size=9, bold=True, color="6B7280")
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        display_val = f"{value}"
        if unit:
            display_val += f" {unit}"
        value_cell.value = display_val
        value_cell.font = Font(name="Calibri", size=16, bold=True, color=color)
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Border around the KPI card
        thin = Side(style='thin', color='E5E7EB')
        card_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # Apply borders and backgrounds
        card_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        for r in range(start_row, start_row + 2):
            for c in range(start_col, start_col + 2):
                cell = worksheet.cell(row=r, column=c)
                cell.fill = card_fill
                
        # Merge values across 2 columns for space
        worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+1)
        worksheet.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+1)

    # --- Single Horizon Exporters ---
    @staticmethod
    def _create_single_summary_sheet(writer, results: dict):
        workbook = writer.book
        worksheet = workbook.create_sheet(title="Summary", index=0)
        worksheet.views.sheetView[0].showGridLines = True
        
        # Title
        worksheet["A1"] = "eDOPT Depot Optimization Summary"
        worksheet["A1"].font = Font(name="Calibri", size=18, bold=True, color="1E3A8A")
        worksheet.row_dimensions[1].height = 40
        
        meta = results.get("run_metadata", {})
        inputs = meta.get("inputs", {})
        settings = meta.get("settings", {})
        co2_factor = inputs.get("co2_emission_factor", 400.0) if inputs else settings.get("co2_emission_factor", 400.0)
        if not co2_factor:
            co2_factor = 400.0
            
        scenario = results.get("scenario", "baseline").upper()
        
        # Section 1: Metadata Table
        worksheet["A3"] = "CONFIGURATION METADATA"
        worksheet["A3"].font = Font(name="Calibri", size=12, bold=True, color="4B5563")
        
        meta_data = [
            ["Scenario Active", scenario],
            ["Time Range", f"{results.get('electricity_prices', [0])*0} {results.get('scenario', 'baseline')} run"],
            ["Planning Horizon", f"{inputs.get('planning_horizon_hours', 24)} Hours"],
            ["Time Step Resolution", f"{inputs.get('time_step_minutes', 15)} min"],
            ["CO2 Emission Factor", f"{co2_factor} g/kWh"],
            ["Depot Charger Capacity", f"{inputs.get('charger_capacity_kw', 1200)} kW"],
            ["Depot Grid Limit", f"{inputs.get('grid_limit_kw', 800)} kW"],
            ["Optimization Run Time", meta.get("timestamp", "N/A")]
        ]
        
        for idx, item in enumerate(meta_data):
            r = 4 + idx
            worksheet.cell(row=r, column=1, value=item[0]).font = Font(bold=True)
            worksheet.cell(row=r, column=2, value=item[1])
            worksheet.cell(row=r, column=1).border = Border(bottom=Side(style='thin', color='E5E7EB'))
            worksheet.cell(row=r, column=2).border = Border(bottom=Side(style='thin', color='E5E7EB'))
            
        # Section 2: Key Metrics Cards
        worksheet["D3"] = "KEY PERFORMANCE INDICATORS"
        worksheet["D3"].font = Font(name="Calibri", size=12, bold=True, color="4B5563")
        
        cost = results.get("total_cost_eur", 0.0)
        base_cost = results.get("baseline_cost_eur", 0.0)
        savings = base_cost - cost if base_cost > 0 else 0.0
        savings_pct = (savings / base_cost * 100) if base_cost > 0 else 0.0
        
        # Calculate carbon emissions
        # Total energy imported
        energy_mwh = results.get("total_energy_mwh", 0.0)
        co2_emissions_t = (energy_mwh * 1000 * co2_factor) / 1e6 # MWh to kWh to tons
        
        # Baseline emissions
        dt = (inputs.get('time_step_minutes', 15) / 60.0) if inputs else 0.25
        base_load_profile = results.get("baseline_aggregated_load_profile", [])
        base_energy_mwh = sum(base_load_profile) * dt / 1000.0 if base_load_profile else energy_mwh
        base_co2_emissions_t = (base_energy_mwh * 1000 * co2_factor) / 1e6
        co2_saved_t = base_co2_emissions_t - co2_emissions_t
        
        ExcelExporter._create_kpi_card(worksheet, start_col=4, start_row=4, label="Total Cost", value=f"€{cost:,.2f}")
        ExcelExporter._create_kpi_card(worksheet, start_col=6, start_row=4, label="Baseline Cost", value=f"€{base_cost:,.2f}", color="6B7280")
        
        ExcelExporter._create_kpi_card(worksheet, start_col=4, start_row=7, label="Financial Savings", value=f"€{savings:,.2f} ({savings_pct:.1f}%)", color="059669" if savings > 0 else "1E3A8A")
        ExcelExporter._create_kpi_card(worksheet, start_col=6, start_row=7, label="Peak Grid Import", value=f"{results.get('peak_load_kw', 0.0):,.1f} kW", color="DC2626")
        
        ExcelExporter._create_kpi_card(worksheet, start_col=4, start_row=10, label="Grid Import Energy", value=f"{energy_mwh:,.2f} MWh")
        ExcelExporter._create_kpi_card(worksheet, start_col=6, start_row=10, label="Estimated CO2 Saved", value=f"{co2_saved_t:.3f} t CO2", color="10B981" if co2_saved_t > 0 else "1E3A8A")

        # Section 3: Scenario PV/BESS breakdown if present
        curr_row = 13
        if "pv" in scenario.lower() or "bess" in scenario.lower():
            worksheet.cell(row=curr_row, column=4, value="RENEWABLES & STORAGE METRICS").font = Font(name="Calibri", size=12, bold=True, color="4B5563")
            curr_row += 1
            
            metrics = []
            if "pv" in scenario.lower():
                metrics.extend([
                    ["Total PV Solar Generated", f"{results.get('total_pv_generated_kwh', 0.0):,.1f} kWh"],
                    ["PV Energy Consumed Locally", f"{results.get('total_pv_used_kwh', 0.0):,.1f} kWh"],
                    ["PV Energy Curtailed", f"{results.get('total_pv_curtailed_kwh', 0.0):,.1f} kWh"]
                ])
            if "bess" in scenario.lower():
                metrics.extend([
                    ["BESS Energy Throughput", f"{results.get('total_bess_throughput_kwh', 0.0):,.1f} kWh"]
                ])
                
            for m in metrics:
                worksheet.cell(row=curr_row, column=4, value=m[0]).font = Font(bold=True)
                worksheet.cell(row=curr_row, column=5, value=m[1])
                worksheet.cell(row=curr_row, column=4).border = Border(bottom=Side(style='thin', color='E5E7EB'))
                worksheet.cell(row=curr_row, column=5).border = Border(bottom=Side(style='thin', color='E5E7EB'))
                curr_row += 1
                
        # Format column widths
        worksheet.column_dimensions["A"].width = 25
        worksheet.column_dimensions["B"].width = 25
        worksheet.column_dimensions["C"].width = 5
        worksheet.column_dimensions["D"].width = 25
        worksheet.column_dimensions["E"].width = 15
        worksheet.column_dimensions["F"].width = 25
        worksheet.column_dimensions["G"].width = 15

    @staticmethod
    def _create_single_timeseries_sheet(writer, results: dict):
        prices = results.get("electricity_prices", [])
        m_prices = results.get("market_prices", [])
        agg_load = results.get("aggregated_load_profile", [])
        base_load = results.get("baseline_aggregated_load_profile", [])
        grid_import = results.get("grid_import_profile", [])
        
        # Build DataFrame
        df_data = {
            "Period Index": list(range(1, len(prices) + 1)),
            "Electricity Price (EUR/MWh)": prices,
        }
        if m_prices:
            df_data["Market Price (EUR/MWh)"] = m_prices
            
        df_data["Optimized Grid Load (kW)"] = agg_load if agg_load else [0]*len(prices)
        if base_load:
            df_data["Baseline Load (kW)"] = base_load
        if grid_import:
            df_data["Net Grid Import (kW)"] = grid_import
            
        # Additional PV
        if results.get("pv_yield_profile"):
            df_data["PV Generation (kW)"] = results.get("pv_yield_profile")
            df_data["PV Consumed (kW)"] = results.get("pv_used_profile", [0]*len(prices))
            df_data["PV Curtailed (kW)"] = results.get("pv_curtailed_profile", [0]*len(prices))
            
        # Additional BESS
        if results.get("bess_soc_profile"):
            # BESS SoC profile size is often N+1, slice to N
            soc_profile = results.get("bess_soc_profile", [])
            if len(soc_profile) > len(prices):
                soc_profile = soc_profile[:len(prices)]
            df_data["BESS Charge Power (kW)"] = results.get("bess_charge_profile", [0]*len(prices))
            df_data["BESS Discharge Power (kW)"] = results.get("bess_discharge_profile", [0]*len(prices))
            df_data["BESS SoC (kWh)"] = soc_profile
            
        # Temperature derating details
        if results.get("ambient_temperature_profile_c"):
            df_data["Ambient Temp (°C)"] = results.get("ambient_temperature_profile_c")
        if results.get("bess_cell_temperature_profile_c"):
            cell_temp = results.get("bess_cell_temperature_profile_c")
            if len(cell_temp) > len(prices):
                cell_temp = cell_temp[:len(prices)]
            df_data["BESS Cell Temp (°C)"] = cell_temp
        if results.get("bess_capacity_factor_profile"):
            df_data["BESS Capacity Factor"] = results.get("bess_capacity_factor_profile")
            
        df = pd.DataFrame(df_data)
        
        # Write to sheet
        sheet_name = "Depot Timeseries"
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ExcelExporter._apply_table_formatting(writer, sheet_name, title="Depot load profiles and timeseries data")

    @staticmethod
    def _create_bus_diagnostics_sheet(writer, results: dict):
        diagnostics = results.get("bus_diagnostics", [])
        if not diagnostics:
            # Create a blank tab
            df = pd.DataFrame([["No diagnostics data available"]], columns=["Message"])
            df.to_excel(writer, sheet_name="Bus Diagnostics", index=False)
            return
            
        rows = []
        for diag in diagnostics:
            # Handle Pydantic models vs standard dicts
            d = diag if isinstance(diag, dict) else diag.dict()
            rows.append([
                d.get("id"),
                d.get("vehicle_type", "Bus"),
                d.get("capacity_kwh", 350.0),
                d.get("initial_soc_kwh", 0.0),
                d.get("final_soc_kwh", 0.0),
                d.get("min_soc_reached_kwh", 0.0),
                d.get("total_trip_energy_kwh", 0.0),
                d.get("total_charged_energy_kwh", 0.0),
                d.get("status_flag", "healthy"),
                d.get("diagnostic_reason", "")
            ])
            
        df = pd.DataFrame(rows, columns=[
            "Bus/Umlauf ID",
            "Vehicle Type",
            "Battery Capacity (kWh)",
            "Initial SoC (kWh)",
            "Final SoC (kWh)",
            "Min SoC Reached (kWh)",
            "Trip Energy (kWh)",
            "Charged Energy (kWh)",
            "Status",
            "Diagnostic Reason"
        ])
        
        sheet_name = "Bus Diagnostics"
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ExcelExporter._apply_table_formatting(writer, sheet_name, title="Fleet Diagnostics & Battery Health")

    @staticmethod
    def _create_bus_schedules_sheet(writer, results: dict):
        schedules = results.get("schedules", {})
        if not schedules:
            df = pd.DataFrame([["No schedule data available"]], columns=["Message"])
            df.to_excel(writer, sheet_name="Bus Schedules (kW)", index=False)
            return
            
        first_bus_schedule = next(iter(schedules.values()))
        num_periods = len(first_bus_schedule)
        
        df_data = {
            "Period Index": list(range(1, num_periods + 1))
        }
        for bus_id, power_list in schedules.items():
            df_data[f"{bus_id} (kW)"] = power_list
            
        df = pd.DataFrame(df_data)
        
        sheet_name = "Bus Schedules (kW)"
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ExcelExporter._apply_table_formatting(writer, sheet_name, title="Individual Bus Charging Power Schedules (kW)")

    @staticmethod
    def _create_bus_soc_sheet(writer, results: dict):
        soc_profiles = results.get("soc_profiles", {})
        if not soc_profiles:
            df = pd.DataFrame([["No SoC profile data available"]], columns=["Message"])
            df.to_excel(writer, sheet_name="Bus SoC (kWh)", index=False)
            return
            
        first_profile = next(iter(soc_profiles.values()))
        num_periods = len(first_profile)
        
        df_data = {
            "Period Index": list(range(1, num_periods + 1))
        }
        for bus_id, soc_list in soc_profiles.items():
            # If the profile is slightly longer or shorter, fit it
            df_data[f"{bus_id} (kWh)"] = soc_list
            
        df = pd.DataFrame(df_data)
        
        sheet_name = "Bus SoC (kWh)"
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ExcelExporter._apply_table_formatting(writer, sheet_name, title="Individual Bus State of Charge Tracking (kWh)")

    # --- Long-Term Exporters ---
    @staticmethod
    def _create_lt_summary_sheet(writer, results: dict):
        workbook = writer.book
        worksheet = workbook.create_sheet(title="Summary", index=0)
        worksheet.views.sheetView[0].showGridLines = True
        
        worksheet["A1"] = "eDOPT Long-Term Simulation Summary"
        worksheet["A1"].font = Font(name="Calibri", size=18, bold=True, color="1E3A8A")
        worksheet.row_dimensions[1].height = 40
        
        meta = results.get("run_metadata", {})
        inputs = meta.get("inputs", {})
        settings = meta.get("settings", {})
        co2_factor = inputs.get("co2_emission_factor", 400.0) if inputs else settings.get("co2_emission_factor", 400.0)
        if not co2_factor:
            co2_factor = 400.0
            
        total_days = results.get("total_days", 1)
        scenario = results.get("scenario", "baseline").upper()
        
        # Meta Table
        worksheet["A3"] = "SIMULATION METADATA"
        worksheet["A3"].font = Font(name="Calibri", size=12, bold=True, color="4B5563")
        
        meta_data = [
            ["Scenario Active", scenario],
            ["Simulation Mode", "Long-Term Simulation"],
            ["Simulation Period", f"{total_days} Days"],
            ["Start Date", inputs.get("start_date", settings.get("longTermConfig", {}).get("startDate", "N/A"))],
            ["End Date", inputs.get("end_date", settings.get("longTermConfig", {}).get("endDate", "N/A"))],
            ["Time Step Resolution", f"{inputs.get('time_step_minutes', 15)} min"],
            ["CO2 Emission Factor", f"{co2_factor} g/kWh"],
            ["Depot Charger Capacity", f"{inputs.get('charger_capacity_kw', 1200)} kW"],
            ["Depot Grid Limit", f"{inputs.get('grid_limit_kw', 800)} kW"],
            ["Optimization Run Time", meta.get("timestamp", "N/A")]
        ]
        
        for idx, item in enumerate(meta_data):
            r = 4 + idx
            worksheet.cell(row=r, column=1, value=item[0]).font = Font(bold=True)
            worksheet.cell(row=r, column=2, value=item[1])
            worksheet.cell(row=r, column=1).border = Border(bottom=Side(style='thin', color='E5E7EB'))
            worksheet.cell(row=r, column=2).border = Border(bottom=Side(style='thin', color='E5E7EB'))
            
        # KPI Cards
        worksheet["D3"] = "LONG-TERM METRICS OVERVIEW"
        worksheet["D3"].font = Font(name="Calibri", size=12, bold=True, color="4B5563")
        
        cost = results.get("total_cost_eur", 0.0)
        base_cost = results.get("baseline_cost_eur", 0.0)
        savings = base_cost - cost if base_cost > 0 else 0.0
        savings_pct = (savings / base_cost * 100) if base_cost > 0 else 0.0
        
        energy_mwh = results.get("total_energy_mwh", 0.0)
        co2_emissions_t = (energy_mwh * 1000 * co2_factor) / 1e6
        
        # Calculate base emissions for long-term
        dt = (inputs.get('time_step_minutes', 15) / 60.0) if inputs else 0.25
        base_load_profile = results.get("baseline_aggregated_load_profile", [])
        base_energy_mwh = sum(base_load_profile) * dt / 1000.0 if base_load_profile else energy_mwh
        base_co2_emissions_t = (base_energy_mwh * 1000 * co2_factor) / 1e6
        co2_saved_t = base_co2_emissions_t - co2_emissions_t
        
        ExcelExporter._create_kpi_card(worksheet, start_col=4, start_row=4, label="Total Cost", value=f"€{cost:,.2f}")
        ExcelExporter._create_kpi_card(worksheet, start_col=6, start_row=4, label="Baseline Cost", value=f"€{base_cost:,.2f}", color="6B7280")
        
        ExcelExporter._create_kpi_card(worksheet, start_col=4, start_row=7, label="Accumulated Savings", value=f"€{savings:,.2f} ({savings_pct:.1f}%)", color="059669" if savings > 0 else "1E3A8A")
        ExcelExporter._create_kpi_card(worksheet, start_col=6, start_row=7, label="Max Peak Load", value=f"{results.get('peak_load_kw', 0.0):,.1f} kW", color="DC2626")
        
        ExcelExporter._create_kpi_card(worksheet, start_col=4, start_row=10, label="Grid Import Energy", value=f"{energy_mwh:,.2f} MWh")
        ExcelExporter._create_kpi_card(worksheet, start_col=6, start_row=10, label="CO2 Offsets Saved", value=f"{co2_saved_t:.3f} t CO2", color="10B981" if co2_saved_t > 0 else "1E3A8A")

        # Solar & BESS Metrics
        curr_row = 13
        if "pv" in scenario.lower() or "bess" in scenario.lower():
            worksheet.cell(row=curr_row, column=4, value="RENEWABLES & STORAGE METRICS").font = Font(name="Calibri", size=12, bold=True, color="4B5563")
            curr_row += 1
            
            metrics = []
            if "pv" in scenario.lower():
                metrics.extend([
                    ["Total PV Solar Generated", f"{results.get('total_pv_generated_kwh', 0.0):,.1f} kWh"],
                    ["PV Energy Consumed Locally", f"{results.get('total_pv_used_kwh', 0.0):,.1f} kWh"],
                    ["PV Energy Curtailed", f"{results.get('total_pv_curtailed_kwh', 0.0):,.1f} kWh"]
                ])
            if "bess" in scenario.lower():
                metrics.extend([
                    ["BESS Energy Throughput", f"{results.get('total_bess_throughput_kwh', 0.0):,.1f} kWh"]
                ])
                
            for m in metrics:
                worksheet.cell(row=curr_row, column=4, value=m[0]).font = Font(bold=True)
                worksheet.cell(row=curr_row, column=5, value=m[1])
                worksheet.cell(row=curr_row, column=4).border = Border(bottom=Side(style='thin', color='E5E7EB'))
                worksheet.cell(row=curr_row, column=5).border = Border(bottom=Side(style='thin', color='E5E7EB'))
                curr_row += 1

        worksheet.column_dimensions["A"].width = 25
        worksheet.column_dimensions["B"].width = 25
        worksheet.column_dimensions["C"].width = 5
        worksheet.column_dimensions["D"].width = 25
        worksheet.column_dimensions["E"].width = 15
        worksheet.column_dimensions["F"].width = 25
        worksheet.column_dimensions["G"].width = 15

    @staticmethod
    def _create_lt_daily_aggregates_sheet(writer, results: dict):
        total_days = results.get("total_days", 1)
        meta = results.get("run_metadata", {})
        inputs = meta.get("inputs", {})
        settings = meta.get("settings", {})
        
        start_date_str = inputs.get("start_date", settings.get("longTermConfig", {}).get("startDate", "2026-01-01"))
        base_date = pd.to_datetime(start_date_str)
        
        # Calculate daily variables from detailed timeseries
        prices = results.get("electricity_prices", [])
        steps_per_day = len(prices) // total_days if total_days > 0 else 96
        dt = (inputs.get('time_step_minutes', 15) / 60.0) if inputs else 0.25
        
        grid_import = results.get("grid_import_profile", [])
        baseline_load = results.get("baseline_aggregated_load_profile", [])
        
        pv_gen = results.get("pv_yield_profile", [])
        bess_charge = results.get("bess_charge_profile", [])
        
        daily_rows = []
        for day in range(total_days):
            date_label = (base_date + pd.Timedelta(days=day)).strftime("%Y-%m-%d")
            start_idx = day * steps_per_day
            end_idx = (day + 1) * steps_per_day
            
            # Slices
            day_prices = prices[start_idx:end_idx]
            day_grid = grid_import[start_idx:end_idx]
            day_base = baseline_load[start_idx:end_idx]
            
            # Calculate Costs (Price is EUR/MWh, Grid import is kW)
            day_cost = sum(g * p / 1000.0 * dt for g, p in zip(day_grid, day_prices)) if day_grid else 0.0
            day_base_cost = sum(b * p / 1000.0 * dt for b, p in zip(day_base, day_prices)) if day_base else day_cost
            day_savings = day_base_cost - day_cost
            
            day_peak = max(day_grid) if day_grid else 0.0
            day_energy = sum(day_grid) * dt / 1000.0 if day_grid else 0.0
            
            row = [
                day + 1,
                date_label,
                round(day_cost, 2),
                round(day_base_cost, 2),
                round(day_savings, 2),
                round(day_peak, 1),
                round(day_energy, 3)
            ]
            
            # PV & BESS daily splits
            if pv_gen:
                day_pv = pv_gen[start_idx:end_idx]
                row.append(round(sum(day_pv) * dt, 1))
            if bess_charge:
                day_bess = bess_charge[start_idx:end_idx]
                row.append(round(sum(day_bess) * dt, 1))
                
            daily_rows.append(row)
            
        cols = [
            "Day Index",
            "Date",
            "Total Cost (EUR)",
            "Baseline Cost (EUR)",
            "Savings (EUR)",
            "Peak Grid Load (kW)",
            "Imported Energy (MWh)"
        ]
        if pv_gen:
            cols.append("Solar PV Generated (kWh)")
        if bess_charge:
            cols.append("BESS Throughput (kWh)")
            
        df = pd.DataFrame(daily_rows, columns=cols)
        
        sheet_name = "Daily Aggregates"
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ExcelExporter._apply_table_formatting(writer, sheet_name, title="Daily Performance Breakdown")

    @staticmethod
    def _create_lt_timeseries_sheet(writer, results: dict):
        # Similar to single horizon, but for the entire range
        ExcelExporter._create_single_timeseries_sheet(writer, results)

    @staticmethod
    def _create_lt_bus_performance_sheet(writer, results: dict):
        diagnostics_by_day = results.get("diagnostics_by_day", [])
        if not diagnostics_by_day:
            df = pd.DataFrame([["No fleet summary data available"]], columns=["Message"])
            df.to_excel(writer, sheet_name="Bus Performance", index=False)
            return
            
        # Accumulate metrics per bus over the multi-day run
        bus_aggregates = {}
        for day_diags in diagnostics_by_day:
            for diag in day_diags:
                d = diag if isinstance(diag, dict) else diag.dict()
                bus_id = d.get("id")
                if bus_id not in bus_aggregates:
                    bus_aggregates[bus_id] = {
                        "id": bus_id,
                        "vehicle_type": d.get("vehicle_type", "Bus"),
                        "capacity_kwh": d.get("capacity_kwh", 350.0),
                        "total_trip_energy": 0.0,
                        "total_charged_energy": 0.0,
                        "min_soc": 9999.0,
                        "soc_samples": [],
                        "critical_days": 0,
                        "warning_days": 0,
                        "healthy_days": 0
                    }
                
                bag = bus_aggregates[bus_id]
                bag["total_trip_energy"] += d.get("total_trip_energy_kwh", 0.0)
                bag["total_charged_energy"] += d.get("total_charged_energy_kwh", 0.0)
                bag["min_soc"] = min(bag["min_soc"], d.get("min_soc_reached_kwh", 0.0))
                bag["soc_samples"].append(d.get("final_soc_kwh", 0.0))
                
                status = d.get("status_flag", "healthy")
                if status == "critical":
                    bag["critical_days"] += 1
                elif status == "warning":
                    bag["warning_days"] += 1
                else:
                    bag["healthy_days"] += 1
                    
        rows = []
        for bag in bus_aggregates.values():
            avg_final_soc = np.mean(bag["soc_samples"]) if bag["soc_samples"] else 0.0
            
            # Determine global status
            if bag["critical_days"] > 0:
                global_status = "critical"
                reason = f"Out-of-energy or limit violation on {bag['critical_days']} days."
            elif bag["warning_days"] > 0:
                global_status = "warning"
                reason = f"Low energy levels reached on {bag['warning_days']} days."
            else:
                global_status = "healthy"
                reason = "Operated within safety limits throughout simulation."
                
            rows.append([
                bag["id"],
                bag["vehicle_type"],
                bag["capacity_kwh"],
                round(bag["total_trip_energy"], 1),
                round(bag["total_charged_energy"], 1),
                round(bag["min_soc"], 1),
                round(avg_final_soc, 1),
                bag["critical_days"] + bag["warning_days"],
                global_status,
                reason
            ])
            
        df = pd.DataFrame(rows, columns=[
            "Bus/Umlauf ID",
            "Vehicle Type",
            "Battery Capacity (kWh)",
            "Total Trip Energy (kWh)",
            "Total Charged Energy (kWh)",
            "Global Minimum SoC (kWh)",
            "Average Final SoC (kWh)",
            "Issues Count (Days)",
            "Global Status",
            "Diagnostic Summary"
        ])
        
        sheet_name = "Bus Performance"
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ExcelExporter._apply_table_formatting(writer, sheet_name, title="Multi-Day Fleet Safety & Energy Performance Summary")
